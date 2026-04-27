"""
qcf_enhanced.py
===============
ENHANCED QCF GENERATOR
Generates side-by-side price comparison Excel

OUTPUT FORMAT:
| Item          | Vendor A | Vendor B | Vendor C | Lowest Price |
|---------------|----------|----------|----------|-------------|
| Steel Rods    | ₹65/kg   | ₹70/kg   | ₹62/kg   | ✅ Vendor C |
| Electric Wire | ₹380/m   | ₹400/m   | No quote | ✅ Vendor A |
"""

import pandas as pd
from datetime import datetime
import os
from typing import Dict, List
import re


def generate_enhanced_qcf(data: Dict, output_dir: str = "data/outputs") -> str:
    """
    Generate enhanced QCF with side-by-side price comparison.
    
    Args:
        data: Dictionary with structure:
            {
                'rfq_subject': str,
                'items': List[Dict],  # RFQ items
                'responses': List[Dict]  # Vendor responses with vendor_name, vendor_email, items
            }
        output_dir: Output directory
    
    Returns:
        Path to generated Excel file
    """
    # Extract data from new format
    rfq_subject = data.get('rfq_subject', 'Quotation Comparison')
    rfq_items = data.get('items', [])
    vendor_responses = data.get('responses', [])
    
    if not vendor_responses:
        print("[⚠] No vendor responses found for comparison")
        return None
    
    # Step 1: Collect all items from all vendors
    all_items = {}  # {item_name: {vendor_name: price}}
    
    for response in vendor_responses:
        vendor_name = response.get('vendor_name', 'Unknown Vendor')
        vendor_items = response.get('items', [])
        
        for item in vendor_items:
            item_name = item.get('item_name', item.get('name', '')).strip()
            price = item.get('price', '').strip()
            
            if item_name:
                if item_name not in all_items:
                    all_items[item_name] = {}
                all_items[item_name][vendor_name] = price if price else 'No quote'
    
    # Step 2: Create comparison table
    comparison_data = []
    
    # Get all unique vendors for column order
    all_vendors = set()
    for vendors in all_items.values():
        all_vendors.update(vendors.keys())
    all_vendors = sorted(all_vendors)
    
    for item_name, vendors in all_items.items():
        row = {"Item": item_name}
        
        # Add each vendor's price (or '-' if no quote)
        for vendor in all_vendors:
            row[vendor] = vendors.get(vendor, '-')
        
        # Find lowest price
        lowest_price = None
        lowest_vendor = None
        
        for vendor, price in vendors.items():
            # Extract numeric value from price string
            numeric_price = _extract_numeric_price(price)
            if numeric_price is not None:
                if lowest_price is None or numeric_price < lowest_price:
                    lowest_price = numeric_price
                    lowest_vendor = vendor
        
        # Add lowest price column
        if lowest_vendor:
            row["Lowest Price"] = f"✅ {lowest_vendor}"
        else:
            row["Lowest Price"] = "-"
        
        comparison_data.append(row)
    
    # Step 3: Create DataFrame
    if not comparison_data:
        print("[⚠] No quotation items found for comparison")
        return None
    
    df = pd.DataFrame(comparison_data)
    
    # Reorder columns: Item, then vendors, then Lowest Price
    column_order = ["Item"] + all_vendors + ["Lowest Price"]
    df = df[column_order]
    
    # Step 4: Generate filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # Sanitize subject for filename
    safe_subject = re.sub(r'[^\w\s-]', '', rfq_subject)[:30]
    safe_subject = re.sub(r'[-\s]+', '_', safe_subject)
    filename = f"qcf_{safe_subject}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Step 5: Write to Excel with formatting
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Price Comparison")
        
        # Get worksheet for formatting
        worksheet = writer.sheets["Price Comparison"]
        
        # Adjust column widths
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            col_letter = chr(64 + idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        # Add title row
        worksheet.insert_rows(1)
        worksheet['A1'] = f"Quotation Comparison: {rfq_subject}"
        worksheet['A1'].font = worksheet['A1'].font.copy(bold=True, size=14)
        
        # Merge title across all columns
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(df.columns))
        worksheet.merge_cells(f'A1:{last_col}1')
    
    print(f"[✓] Enhanced QCF generated: {filepath}")
    return filepath


def _extract_numeric_price(price_str: str) -> float:
    """
    Extract numeric value from price string.
    
    Examples:
        "₹65" -> 65.0
        "Rs. 1,200" -> 1200.0
        "$100.50" -> 100.5
    """
    try:
        # Remove currency symbols and spaces
        cleaned = re.sub(r'[₹$£€Rs\.,\s]', '', price_str)
        
        # Extract first number
        match = re.search(r'\d+\.?\d*', cleaned)
        if match:
            return float(match.group())
        return None
    except:
        return None


def generate_summary_report(responses: Dict) -> str:
    """
    Generate text summary of all responses.
    
    Args:
        responses: Dictionary of vendor responses (OLD FORMAT - for backward compatibility)
    
    Returns:
        Formatted text report
    """
    report = "\n" + "="*70 + "\n"
    report += "VENDOR QUOTATION SUMMARY\n"
    report += "="*70 + "\n\n"
    
    # Statistics
    total = len(responses)
    quotations = sum(1 for r in responses.values() 
                     if r.get("parsed_data", {}).get("is_quotation"))
    
    report += f"📊 Statistics:\n"
    report += f"   Total Vendors: {total}\n"
    report += f"   Quotations Received: {quotations}\n"
    report += f"   Response Rate: {quotations/total*100:.1f}%\n\n"
    
    # Detailed list
    report += "📋 Detailed Responses:\n"
    report += "-"*70 + "\n"
    
    for email, resp in responses.items():
        vendor_name = resp.get("name", "Unknown")
        status = resp.get("status", "Pending")
        
        report += f"\n• {vendor_name} ({email})\n"
        report += f"  Status: {status}\n"
        
        parsed = resp.get("parsed_data")
        if parsed and parsed.get("is_quotation"):
            items = parsed.get("items", [])
            report += f"  Items: {len(items)}\n"
            
            for item in items:
                report += f"    - {item.get('item_name')}: {item.get('price')}\n"
        
        report += "-"*70 + "\n"
    
    return report


if __name__ == "__main__":
    print("Enhanced QCF Generator Module Loaded Successfully!")
